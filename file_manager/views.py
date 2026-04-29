from django.shortcuts import redirect, render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db.models import Count, ExpressionWrapper, IntegerField
from .models import Folder, SharedFile, PasswordResetRequest
from django.views.decorators.cache import never_cache
from django.contrib.auth.models import User
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
from collections import defaultdict
from django.utils import timezone
from datetime import datetime, timedelta
from django.http import JsonResponse
from django.views.decorators.http import require_POST
import json
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def parse_datetime_local(value: str):
    """
    Parse HTML `datetime-local` value (e.g. '2026-04-29T13:45') into
    an aware datetime in the server's current timezone.
    """
    value = (value or "").strip()
    if not value:
        return None

    try:
        dt = datetime.fromisoformat(value)
    except ValueError:
        return None

    if timezone.is_naive(dt):
        return timezone.make_aware(dt, timezone.get_current_timezone())
    return dt

@login_required
@staff_member_required
def edit_folder(request, folder_id):
    folder = get_object_or_404(Folder, id=folder_id, master_user=request.user)
    
    if request.method == "POST":
        new_name = request.POST.get("folder_name").strip()
        
        # Check agar is naam ka folder pehle se toh nahi hai (khud ko exclude karke)
        if Folder.objects.filter(
            name__iexact=new_name,
            master_user=request.user,
            parent=folder.parent,
        ).exclude(id=folder_id).exists():
            messages.error(request, f'Folder "{new_name}" already exists!')
        elif new_name:
            folder.name = new_name
            folder.save()
            messages.success(request, f'Folder renamed to "{new_name}" successfully!')
            
    return redirect('folder_detail', folder_id=folder.id)

@login_required
@staff_member_required
def delete_folder(request, folder_id):
    folder = get_object_or_404(Folder, id=folder_id, master_user=request.user)
    parent_id = folder.parent_id
    
    if request.method == "POST":
        folder_name = folder.name
        folder.delete()
        messages.success(request, f'Folder "{folder_name}" and its contents deleted.')

    # After delete, go back to parent folder if it exists
    if parent_id:
        return redirect('folder_detail', folder_id=parent_id)
    return redirect('dashboard')


@login_required
@staff_member_required
@never_cache
def folder_detail(request, folder_id):
    folder = get_object_or_404(Folder, id=folder_id, master_user=request.user)

    # Handle create sub-folder + upload inside this folder
    if request.method == "POST":
        if "create_folder" in request.POST:
            new_name = request.POST.get("folder_name", "").strip()
            if not new_name:
                return redirect('folder_detail', folder_id=folder.id)

            if Folder.objects.filter(
                master_user=request.user,
                parent=folder,
                name__iexact=new_name
            ).exists():
                messages.error(request, f'Folder "{new_name}" already exists!')
                return redirect('folder_detail', folder_id=folder.id)

            new_folder = Folder.objects.create(
                name=new_name,
                master_user=request.user,
                parent=folder
            )
            messages.success(request, f'Folder "{new_name}" created successfully!')
            # Redirect into newly created folder so user can keep creating inside it
            return redirect('folder_detail', folder_id=new_folder.id)

        if "upload_file" in request.POST:
            title = request.POST.get("title", "").strip()
            file_obj = request.FILES.get("file")
            expiry_at_str = request.POST.get("expiry_at", "").strip()
            expiry_at = parse_datetime_local(expiry_at_str) if expiry_at_str else None
            duration = request.POST.get("duration", 2)
            user_ids = request.POST.getlist("users")

            if title and file_obj:
                if expiry_at:
                    if expiry_at <= timezone.now():
                        messages.error(request, "Expiry time must be in the future.")
                        return redirect('folder_detail', folder_id=folder.id)
                    new_file = SharedFile.objects.create(
                        title=title,
                        file=file_obj,
                        master_user=request.user,
                        folder=folder,
                        expiry_at=expiry_at,
                        duration_hrs=0,
                    )
                else:
                    new_file = SharedFile.objects.create(
                        title=title,
                        file=file_obj,
                        master_user=request.user,
                        folder=folder,
                        duration_hrs=int(duration),
                    )
                new_file.allowed_users.set(user_ids)
                messages.success(request, f'File "{title}" uploaded and shared.')
            return redirect('folder_detail', folder_id=folder.id)

    # Data fetching
    subfolders = Folder.objects.filter(master_user=request.user, parent=folder)
    folder_files = SharedFile.objects.filter(master_user=request.user, folder=folder)
    clients = User.objects.filter(is_staff=False)

    return render(request, 'file_manager/folder_detail.html', {
        'folder': folder,
        'subfolders': subfolders,
        'folder_files': folder_files,
        'clients': clients,
    })

@login_required(login_url='login')
@never_cache
def user_dashboard(request):
    if request.user.is_staff:
        # --- FOLDER LOGIC ---
        if request.method == "POST" and "create_folder" in request.POST:
            name = request.POST.get("folder_name").strip()
            
            # Case-insensitive duplicate check
            if Folder.objects.filter(
                name__iexact=name,
                master_user=request.user,
                parent__isnull=True
            ).exists():
                messages.error(request, f'Folder "{name}" already exists!')
                return redirect('dashboard') 
            
            elif name:
                new_folder = Folder.objects.create(name=name, master_user=request.user, parent=None)
                messages.success(request, f'Folder "{name}" created successfully!')
                return redirect('folder_detail', folder_id=new_folder.id)

        # --- FILE UPLOAD LOGIC ---
        if request.method == "POST" and "upload_file" in request.POST:
            title = request.POST.get("title")
            file_obj = request.FILES.get("file")
            folder_id = request.POST.get("folder")
            expiry_at_str = request.POST.get("expiry_at", "").strip()
            expiry_at = parse_datetime_local(expiry_at_str) if expiry_at_str else None
            duration = request.POST.get("duration", 2)
            user_ids = request.POST.getlist("users")

            if expiry_at:
                if expiry_at <= timezone.now():
                    messages.error(request, "Expiry time must be in the future.")
                    return redirect('dashboard')
                new_file = SharedFile.objects.create(
                    title=title,
                    file=file_obj,
                    master_user=request.user,
                    folder_id=folder_id if folder_id else None,
                    expiry_at=expiry_at,
                    duration_hrs=0,
                )
            else:
                new_file = SharedFile.objects.create(
                    title=title,
                    file=file_obj,
                    master_user=request.user,
                    folder_id=folder_id if folder_id else None,
                    duration_hrs=int(duration)
                )
            new_file.allowed_users.set(user_ids)
            messages.success(request, f'File "{title}" uploaded and shared.')
            return redirect('dashboard')

        # Data Fetching
        folders = Folder.objects.filter(
            master_user=request.user,
            parent__isnull=True,
        ).annotate(
            items_count=ExpressionWrapper(
                Count('children', distinct=True) + Count('files', distinct=True),
                output_field=IntegerField(),
            )
        )
        # Master should be able to see all shared files, including those inside folders.
        root_files = SharedFile.objects.filter(master_user=request.user).select_related('folder')
        clients = User.objects.filter(is_staff=False)
        
        return render(request, 'file_manager/master_dashboard.html', {
            'folders': folders,
            'root_files': root_files,
            'clients': clients
        })
    
    else:
        # Client view
        files = SharedFile.objects.filter(allowed_users=request.user).distinct()
        has_pending_reset = PasswordResetRequest.objects.filter(
            requested_user=request.user,
            status=PasswordResetRequest.STATUS_PENDING
        ).exists()
        return render(request, 'file_manager/dashboard.html', {
            'files': files,
            'has_pending_reset': has_pending_reset,
        })


@login_required
@staff_member_required
@never_cache
def manage_clients(request):
    if request.method == "POST":
        if "create_client" in request.POST:
            username = request.POST.get("username", "").strip()
            email = request.POST.get("email", "").strip()
            password = request.POST.get("password", "").strip()

            if not username or not password:
                messages.error(request, "Username and password are required.")
                return redirect('manage_clients')

            if User.objects.filter(username__iexact=username).exists():
                messages.error(request, f'Client "{username}" already exists!')
                return redirect('manage_clients')

            new_client = User.objects.create_user(
                username=username,
                email=email,
                password=password,
            )
            new_client.is_staff = False
            new_client.save()
            messages.success(request, f'Client "{username}" created successfully!')
            return redirect('manage_clients')

        if "delete_client" in request.POST:
            client_id = request.POST.get("client_id")
            if client_id:
                client = get_object_or_404(User, id=client_id, is_staff=False)
                client.delete()
                messages.success(request, "Client deleted successfully.")
            return redirect('manage_clients')

        if "change_password" in request.POST:
            client_id = request.POST.get("client_id")
            new_password = request.POST.get("new_password", "").strip()
            if client_id and new_password:
                client = get_object_or_404(User, id=client_id, is_staff=False)
                client.set_password(new_password)
                client.save()
                messages.success(request, f'Password for "{client.username}" updated successfully.')
            else:
                messages.error(request, "Password cannot be empty.")
            return redirect('manage_clients')

    clients = User.objects.filter(is_staff=False).order_by('username')
    return render(request, 'file_manager/manage_clients.html', {'clients': clients})


@login_required
@staff_member_required
@never_cache
def all_folders(request):
    folders = Folder.objects.filter(master_user=request.user).annotate(
        direct_subfolders=Count('children', distinct=True),
        direct_files=Count('files', distinct=True),
    )

    children_map = defaultdict(list)
    for f in folders:
        children_map[f.parent_id].append(f)

    # Sort children for stable UI ordering
    for parent_id in list(children_map.keys()):
        children_map[parent_id].sort(key=lambda x: (x.name or "").lower())

    flat_folders = []

    def dfs(node, depth):
        flat_folders.append({
            'folder': node,
            'depth': depth,
            'indent_px': depth * 20,
            'items_count': (node.direct_subfolders or 0) + (node.direct_files or 0),
        })
        for child in children_map.get(node.id, []):
            dfs(child, depth + 1)

    for root in children_map.get(None, []):
        dfs(root, 0)

    return render(request, 'file_manager/all_folders.html', {'flat_folders': flat_folders})

@login_required
@never_cache
def file_detail(request, file_id):
    if request.user.is_staff:
        file_obj = get_object_or_404(SharedFile, id=file_id, master_user=request.user)
    else:
        file_obj = get_object_or_404(SharedFile, id=file_id, allowed_users=request.user)

    # Master can replace an active file (within duration window).
    if request.method == "POST" and request.user.is_staff:
        if not file_obj.is_editable:
            messages.error(request, "This file is now in View Only mode. Editing is disabled.")
            return redirect('file_detail', file_id=file_obj.id)

        new_file_obj = request.FILES.get("new_file")
        if not new_file_obj:
            messages.error(request, "Please select a file to upload.")
            return redirect('file_detail', file_id=file_obj.id)

        expiry_at_str = request.POST.get("expiry_at", "").strip()
        expiry_at = parse_datetime_local(expiry_at_str) if expiry_at_str else None
        if expiry_at and expiry_at <= timezone.now():
            messages.error(request, "Expiry time must be in the future.")
            return redirect('file_detail', file_id=file_obj.id)

        file_obj.file = new_file_obj
        file_obj.upload_time = timezone.now()  # reset expiry window
        if expiry_at:
            file_obj.expiry_at = expiry_at
        file_obj.save()
        messages.success(request, "File updated successfully.")
        return redirect('file_detail', file_id=file_obj.id)

    ext = file_obj.file.name.lower()
    is_excel = ext.endswith('.xlsx') or ext.endswith('.xls')
    is_pdf   = ext.endswith('.pdf')
    is_docx  = ext.endswith('.docx') or ext.endswith('.doc')

    return render(request, 'file_manager/file_detail.html', {
        'file': file_obj,
        'is_excel': is_excel,
        'is_pdf': is_pdf,
        'is_docx': is_docx,
    })


def _excel_to_luckysheet(file_path):
    """Convert an xlsx file to a Luckysheet-compatible celldata list."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        celldata = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                ct = {
                    "r": cell.row - 1,
                    "c": cell.column - 1,
                    "v": {"v": cell.value, "m": str(cell.value)},
                }
                if cell.font and cell.font.bold:
                    ct["v"]["bl"] = 1
                if cell.font and cell.font.italic:
                    ct["v"]["it"] = 1
                celldata.append(ct)
        sheets.append({
            "name": ws.title,
            "celldata": celldata,
            "row": max(ws.max_row or 1, 100),
            "column": max(ws.max_column or 1, 26),
        })
    return sheets


@login_required
@require_POST
def save_excel(request, file_id):
    """Receive Luckysheet JSON and write it back as .xlsx"""
    if request.user.is_staff:
        file_obj = get_object_or_404(SharedFile, id=file_id, master_user=request.user)
    else:
        file_obj = get_object_or_404(SharedFile, id=file_id, allowed_users=request.user)
        if not file_obj.is_editable:
            return JsonResponse({"error": "File has expired. Editing is disabled."}, status=403)

    try:
        payload = json.loads(request.body)
        sheets_data = payload.get("sheets", [])
    except (json.JSONDecodeError, KeyError):
        return JsonResponse({"error": "Invalid data."}, status=400)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet in sheets_data:
        ws = wb.create_sheet(title=sheet.get("name", "Sheet"))
        for cell_entry in sheet.get("celldata", []):
            r = cell_entry.get("r", 0) + 1
            c = cell_entry.get("c", 0) + 1
            v = cell_entry.get("v", {})

            # v can be a plain scalar (string/number) or a Luckysheet cell object.
            # Luckysheet cell objects look like: {"v": 42, "m": "42", "ct": {...}}
            if isinstance(v, dict):
                value = v.get("v", "")
                bold   = bool(v.get("bl", 0))
                italic = bool(v.get("it", 0))
            else:
                # Plain scalar — came from a simple cell
                value  = v
                bold   = False
                italic = False

            if value is None or value == "":
                continue

            ws.cell(row=r, column=c, value=value)
            if bold or italic:
                ws.cell(row=r, column=c).font = Font(bold=bold, italic=italic)

    wb.save(file_obj.file.path)
    return JsonResponse({"status": "saved"})


@login_required
def load_excel(request, file_id):
    """Return Excel sheet data as JSON for Luckysheet."""
    if request.user.is_staff:
        file_obj = get_object_or_404(SharedFile, id=file_id, master_user=request.user)
    else:
        file_obj = get_object_or_404(SharedFile, id=file_id, allowed_users=request.user)

    try:
        sheets = _excel_to_luckysheet(file_obj.file.path)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"sheets": sheets})


@login_required
@staff_member_required
def delete_shared_file(request, file_id):
    shared_file = get_object_or_404(SharedFile, id=file_id, master_user=request.user)

    if request.method == "POST":
        title = shared_file.title
        # Remove the physical file from storage (best-effort).
        try:
            shared_file.file.delete(save=False)
        except Exception:
            pass
        shared_file.delete()
        messages.success(request, f'File "{title}" deleted successfully.')

    return redirect('dashboard')


@login_required
@staff_member_required
@never_cache
def master_password_reset_requests(request):
    # Staff can see all requests; most recent first.
    requests_qs = PasswordResetRequest.objects.select_related('requested_user').all()

    return render(request, 'file_manager/password_reset_requests.html', {
        'requests': requests_qs,
        'temp_password': None,
    })


@login_required
def request_password_reset(request):
    # Only clients (non-staff) should be allowed to request resets.
    if request.user.is_staff:
        return redirect('dashboard')

    if request.method == "POST":
        already_pending = PasswordResetRequest.objects.filter(
            requested_user=request.user,
            status=PasswordResetRequest.STATUS_PENDING,
        ).exists()

        if already_pending:
            messages.info(request, "You already requested a password reset. Please wait for master approval.")
            return redirect('dashboard')

        PasswordResetRequest.objects.create(
            requested_user=request.user,
            status=PasswordResetRequest.STATUS_PENDING,
        )
        messages.success(request, "Password reset requested. Master will review and update it.")
        return redirect('dashboard')

    # If someone visits GET, just redirect back.
    return redirect('dashboard')


@login_required
@staff_member_required
def master_reset_password(request, req_id):
    reset_request = get_object_or_404(PasswordResetRequest, id=req_id)

    if reset_request.status != PasswordResetRequest.STATUS_PENDING:
        messages.info(request, "This request is already handled.")
        return redirect('master_password_reset_requests')

    temp_password = None

    if request.method == "POST":
        temp_password = PasswordResetRequest.generate_temporary_password(14)
        user_obj = reset_request.requested_user
        user_obj.set_password(temp_password)
        user_obj.save()

        reset_request.status = PasswordResetRequest.STATUS_APPROVED
        reset_request.executed_at = timezone.now()
        reset_request.save()

        messages.success(request, "Password reset completed. Copy the temporary password and share it with the user.")

    # Render without redirect so master can see the generated password.
    return render(request, 'file_manager/password_reset_requests.html', {
        'requests': PasswordResetRequest.objects.select_related('requested_user').all(),
        'temp_password': temp_password,
        'handled_request': reset_request,
    })


@login_required
@staff_member_required
def master_inform_password_reset(request, req_id):
    reset_request = get_object_or_404(PasswordResetRequest, id=req_id)

    if reset_request.status != PasswordResetRequest.STATUS_PENDING:
        messages.info(request, "This request is already handled.")
        return redirect('master_password_reset_requests')

    if request.method == "POST":
        note = request.POST.get('master_note', '').strip()
        reset_request.status = PasswordResetRequest.STATUS_INFORMED
        reset_request.master_note = note
        reset_request.executed_at = timezone.now()
        reset_request.save()
        messages.success(request, "Marked as informed. You can now tell the user manually.")

    return redirect('master_password_reset_requests')